import openpyxl


class HomePageData:
    test_HomePage_data =  ([{"firstname": "Sai", "lastname": "Kumar", "gender": "Male"}, ("Sai", "Kumar", "Male"), ("Sai", "Pallavi", "Female")])

    @staticmethod
    def getTestData(self,test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\Sai Kumar\OneDrive\Documents\excelDemo - Copy.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row = i, column = 1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="Rolex
                    Dict[sheet.cell(row = 1, column = j).value] = sheet.cell(row = i, column = j).value

        return[Dict]
